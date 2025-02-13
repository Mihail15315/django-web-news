from django.shortcuts import render, get_object_or_404
import requests
from rest_framework import viewsets
from .models import News
from .serializers import NewsSerializer
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from .forms import UploadFileForm
from django.contrib.gis.geos import Point
import openpyxl
import xlsxwriter
from .models import WeatherReport
import logging
from datetime import datetime

@csrf_exempt
def index(request):
    if request.method == 'POST':
        message = request.POST['message']
        email = request.POST['email']
        name = request.POST['name']
        send_mail(
        name,
        message,
        settings.EMAIL_HOST_USER,
        ['mihailbaranov780@gmail.com'],
        fail_silently=False,
        )
        return HttpResponse("Message sent successfully!")

class NewsViewSet(viewsets.ModelViewSet):
    queryset = News.objects.all()
    serializer_class = NewsSerializer
def news_list(request):
    response = requests.get('http://127.0.0.1:8000/api/v1/news/')
    news_items = response.json() if response.status_code == 200 else []
    return render(request, 'news/news_list.html', {'news_items': news_items})

def news_detail(request, news_id):
    news_item = get_object_or_404(News, id=news_id)
    return render(request, 'news/news_detail.html', {'news_item': news_item})

from django.shortcuts import redirect
from .forms import NotablePlaceForm
from .models import NotablePlace

def edit_place(request, pk):
    place = get_object_or_404(NotablePlace, pk=pk)
    if request.method == 'POST':
        form = NotablePlaceForm(request.POST, instance=place)
        if form.is_valid():
            form.save()
            return redirect('http://127.0.0.1:8000/admin/news/notableplace/')
    else:
        form = NotablePlaceForm(instance=place)
    return render(request, 'news/notableplace/change_form.html', {
    'form': form,
    'original': place,  # Передаем объект, который редактируется
    'opts': NotablePlace._meta,  # Передаем метаданные модели
})
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                name, coordinates, rating = row
                latitude, longitude = map(float, coordinates.split(','))
                point = Point(longitude, latitude)
                NotablePlace.objects.create(
                    name=name,
                    coordinates=point,
                    rating=rating,
                )
            return render(request, 'places/upload_success.html')
    else:
        form = UploadFileForm()
    return render(request, 'places/upload.html', {'form': form})

def export_weather_report(request):
    places = NotablePlace.objects.all()

    if request.method == 'GET':
        place_id = request.GET.get('place_id')  # Используем place_id вместо place
        date_str = request.GET.get('date')

        # Фильтрация данных
        weather_reports = WeatherReport.objects.all()

        if place_id:
            print(f"Filtering by place_id: {place_id}")  # Отладочное сообщение
            weather_reports = weather_reports.filter(place_id=place_id)

        if date_str:
            print(f"Filtering by date: {date_str}")  # Отладочное сообщение
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                weather_reports = weather_reports.filter(timestamp__date=date)
            except ValueError:
                print("Invalid date format. Expected YYYY-MM-DD.")

        # Создание xlsx-файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="weather_reports.xlsx"'

        workbook = xlsxwriter.Workbook(response, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Заголовки столбцов
        headers = ['Место', 'Температура (°C)', 'Влажность (%)', 'Давление (hPa)', 'Направление ветра', 'Скорость ветра (м/с)', 'Дата и время']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header)

        # Данные
        for row_num, report in enumerate(weather_reports, start=1):
            worksheet.write(row_num, 0, report.place.name)
            worksheet.write(row_num, 1, report.temperature)
            worksheet.write(row_num, 2, report.humidity)
            worksheet.write(row_num, 3, report.pressure)
            worksheet.write(row_num, 4, report.wind_direction)
            worksheet.write(row_num, 5, report.wind_speed)
            worksheet.write(row_num, 6, report.timestamp.strftime('%Y-%m-%d %H:%M:%S'))

        workbook.close()
        return response

    # Если запрос не GET, отображаем HTML-страницу
    return render(request, 'news/export_weather_reports.html', {'places': places})